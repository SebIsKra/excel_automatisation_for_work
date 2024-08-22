import os
import openpyxl
import pandas as pd 

# Function to remove write protection from an Excel file
def remove_write_protection(excel_file):
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=False)

        # Remove write protection (if any)
        wb.security = None

        # Save the changes
        wb.save(excel_file)
    except Exception as e:
        print(f"Error removing write protection from '{excel_file}': {e}")

#Function to append rows 2 and 3 from one Excel file to another
def append_rows(source_file, target_file):
    try:
        #Open the excel to append
        appended_rows_excel = openpyxl.load_workbook(source_file, read_only=True)
        main_excel = openpyxl.load_workbook(target_file)
        
        appended_rows_ws = appended_rows_excel.active
        main_ws = main_excel.active

        # Append rows 2 and 3 from source to target
        for row in appended_rows_ws.iter_rows(min_row=2, max_row=3):
            main_ws.append([cell.value for cell in row])

        main_excel.save(target_file)
    except Exception as e:
        print(f"Error appending rows from '{source_file}' to '{target_file}': {e}")

# Function to copy the content of an Excel file to the clipboard
def copy_to_clipboard(excel_file):
    try:
        #Defines the columns of the facebook import excel
        necessary_columns=["Campaign ID", "Campaign Name", "Campaign Status", "Campaign Objective", "Buying Type",	"Ad Set ID", "Ad Set Run Status", "Ad Set Name", "Ad Set Time Start", "Ad Set Time Stop", "Ad Set Daily Budget", "Ad Set Lifetime Budget",	"Link Object ID",	"Link",	"Application ID", "Countries", "Global Regions", "Excluded Global Regions",	"Locales",	"Cities",	"Regions",	"Zip",	"Gender",	"Age Min",	"Age Max",	"Education Status",	"Interested In",	"Relationship",	"Connections",	"Excluded Connections",	"Friends of Connections",	"Broad Category Clusters",	"Custom Audiences",	"Excluded Custom Audiences",	"Publisher Platforms",	"Device Platforms",	"Facebook Positions",	"Instagram Positions",	"Messenger Positions",	"Audience Network Positions",	"Optimization Goal",	"Billing Event",	"Bid Amount",	"Ad ID",	"Ad Status",	"Ad Name",	"Title",	"Body",	"Link Description",	"Display Link",	"Conversion Tracking Pixels",	"Creative Type",	"Creative Optimization",	"Image",	"Image Hash",	"Call to Action",	"Story ID"]
        
        #Turn the excel into a dataframe for pandas using the specified columns
        df = pd.read_excel(excel_file, usecols=necessary_columns, dtype=str)
        #Replaces the dot in the numerical data in "Bid Amount" with comma so the format fits the meta import
        df["Bid Amount"] = df["Bid Amount"].str.replace('.', ',')
        #Copies content of the dataframe into the excel file
        df.to_clipboard(excel=True, index=False, sep="\t")
    except Exception as e:
        print(f"Error copying content of \"{excel_file}\" to clipboard: {e}")

# Function to check if a file is a valid Excel file
def is_valid_excel_file(file_path):
    try:
        openpyxl.load_workbook(file_path, read_only=True)
        return True
    except Exception:
        return False


def main():
    # Get the path to the user's home download folder 
    downloads_folder = os.path.expanduser('~/Downloads')

    # List all files in the Downloads folder
    files = os.listdir(downloads_folder)

    # Find the first valid Excel file in the Downloads folder
    first_excel_file = None
    processed_files = set()

    
    for file in files:
        if file.startswith('picture_facebook'):
            full_path = os.path.join(downloads_folder, file)
            if is_valid_excel_file(full_path):
                if first_excel_file is None:
                    first_excel_file = full_path
                    print("First valid Excel file found:", first_excel_file)

                    # Remove write protection from the first Excel file and add to processed_files list
                    remove_write_protection(first_excel_file)
                    processed_files.add(full_path)
                else:
                    # Append rows from other valid Excel files to the first Excel file
                    if full_path not in processed_files:
                        append_rows(full_path, first_excel_file)
                        processed_files.add(full_path)

    if first_excel_file:
        print("Rows appended from other Excel files to the first Excel file.")

        # Copy the content of the first Excel file to the clipboard
        copy_to_clipboard(first_excel_file)

        print("Content of the first Excel file copied to the clipboard.")
    else:
        print("No valid Excel file found in the Downloads folder.")

    # Print out the processed files
    print("Processed files:")
    for file in processed_files:
        print(file)

if __name__ == "__main__":
    main()